import os
import numpy as np
from ultralytics import YOLO
from PIL import Image
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== CONFIGURE THESE =====
ROOT_FOLDER  = r"C:\Users\kskad\Downloads\clear"
MODEL        = "yolo11n.pt"
CONFIDENCE   = 0.3
OUTPUT_EXCEL = "yolo_results.xlsx"
# ===========================

# Folder layout:
#   ROOT_FOLDER/
#     low/   d10/, d15/, d20/   <- 320x240
#     mid/   d10/, d15/, d20/   <- 800x600
#     high/  d10/, d15/, d20/   <- 1024x768
#
# Files inside each distance folder are matched by checking that the
# time-slot string (e.g. "20_00") appears in the filename. Tolerates
# minor naming variations.

PERSON_CLASS = 0
DISTANCES    = ["d10", "d15", "d20"]
DIST_LABELS  = {"d10": "10m", "d15": "15m", "d20": "20m"}
RESOLUTIONS  = ["low", "mid", "high"]
TIME_SLOTS   = ["20_45", "20_50", "20_55", "20_57", "21_00", "21_02", "21_05", "21_07", "21_10", "21_12", "21_15", "21_17", "21_20", "21_22", "21_25"]
TIME_LABELS  = ["20:45", "20:50", "20:55", "20:57", "21:00", "21:02", "21:05", "21:07", "21:10", "21:12", "21:15", "21:17", "21:20", "21:22", "21:25"]
RES_MAP      = {"low": "320x240", "mid": "800x600", "high": "1024x768"}

print("Loading YOLO model...")
model = YOLO(MODEL)

# ---------------------------------------------------------------
# Walk folders: ROOT/<res>/<dist>/<files>
# ---------------------------------------------------------------
print("Scanning folder structure...")
folder_files = {}     # (res, dist) -> list of (lowercase_name, full_path)
missing_folders = []

for res in RESOLUTIONS:
    for dist in DISTANCES:
        sub = os.path.join(ROOT_FOLDER, res, dist)
        if not os.path.isdir(sub):
            missing_folders.append(sub)
            folder_files[(res, dist)] = []
            continue
        files = []
        for f in os.listdir(sub):
            if f.lower().endswith((".jpg", ".jpeg")):
                files.append((f.lower(), os.path.join(sub, f)))
        folder_files[(res, dist)] = files

total = sum(len(v) for v in folder_files.values())
print(f"Found {total} images across {len(folder_files)} folders.")
if missing_folders:
    print("Missing folders (will be skipped):")
    for m in missing_folders:
        print(f"  {m}")

# ---------------------------------------------------------------
# Run detection
# ---------------------------------------------------------------
results = {}
missing = []

for dist in DISTANCES:
    results[dist] = {}
    for res in RESOLUTIONS:
        results[dist][res] = {}

        for time in TIME_SLOTS:
            candidates = [
                path for (lname, path) in folder_files[(res, dist)]
                if time in lname
            ]

            if not candidates:
                label = f"{res}/{dist}/{time}"
                print(f"  Missing: {label}")
                missing.append(label)
                results[dist][res][time] = None
                continue

            filepath = sorted(candidates)[0]

            img_pil = Image.open(filepath).convert("RGB")
            img_np  = np.array(img_pil)

            det = model.predict(
                source  = img_np,
                classes = [PERSON_CLASS],
                conf    = CONFIDENCE,
                verbose = False
            )[0]

            confs = det.boxes.conf.tolist() if len(det.boxes) > 0 else []
            score = round(max(confs), 2) if confs else 0.0
            results[dist][res][time] = score

            status = (f"  {dist}/{res}/{time}: {score:.2f}"
                      if score > 0 else
                      f"  {dist}/{res}/{time}: not detected")
            print(status)

# ---------------------------------------------------------------
# Styles
# ---------------------------------------------------------------
def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

FONT_TITLE  = Font(name="Arial", bold=True, size=13)
FONT_HEADER = Font(name="Arial", bold=True, size=10)
FONT_DIST   = Font(name="Arial", bold=True, size=10)
FONT_RES    = Font(name="Arial", size=9, italic=True)
FONT_DATA   = Font(name="Arial", size=10)
FONT_NOTE   = Font(name="Arial", italic=True, size=9)

# ---------------------------------------------------------------
# Build Excel (single sheet now)
# ---------------------------------------------------------------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Clear Background"

n_times    = len(TIME_SLOTS)
last_col_i = 2 + n_times                       # Distance, Resolution, plus times
last_col   = get_column_letter(last_col_i)

# Column widths
ws.column_dimensions["A"].width = 10
ws.column_dimensions["B"].width = 18
for i in range(n_times):
    ws.column_dimensions[get_column_letter(3 + i)].width = 10

# Row 1: Title
ws.merge_cells(f"A1:{last_col}1")
c           = ws["A1"]
c.value     = "YOLOv11 Person Detection Results — Clear Background"
c.font      = FONT_TITLE
c.alignment = CENTER
ws.row_dimensions[1].height = 28

# Row 2: Subtitle
ws.merge_cells(f"A2:{last_col}2")
c           = ws["A2"]
c.value     = ("YOLOv11n confidence scores (0.00–1.00) for pedestrian detection. "
               "Score of 0.00 indicates no detection. Evaluated across three distances, "
               "three resolutions, and multiple capture times.")
c.font      = FONT_NOTE
c.alignment = LEFT
ws.row_dimensions[2].height = 30

# Row 3: Column headers
ws.row_dimensions[3].height = 30
for col, label in enumerate(["Distance", "Resolution"] + TIME_LABELS, start=1):
    c           = ws.cell(row=3, column=col)
    c.value     = label
    c.font      = FONT_HEADER
    c.alignment = CENTER
    c.border    = thin_border()

# Data rows
row = 4

for dist in DISTANCES:
    dist_start = row

    for res in RESOLUTIONS:
        res_label = f"{res.capitalize()} ({RES_MAP[res]})"

        c           = ws.cell(row=row, column=2)
        c.value     = res_label
        c.font      = FONT_RES
        c.alignment = CENTER
        c.border    = thin_border()

        for t_idx, time in enumerate(TIME_SLOTS):
            val = results[dist][res][time]
            c   = ws.cell(row=row, column=3 + t_idx)

            if val is None:
                c.value = "—"
            elif val == 0.0:
                c.value = "0.00"
            else:
                c.value = f"{val:.2f}"

            c.font      = FONT_DATA
            c.alignment = CENTER
            c.border    = thin_border()

        ws.row_dimensions[row].height = 20
        row += 1

    # Distance cell
    ws.merge_cells(f"A{dist_start}:A{row - 1}")
    dc           = ws.cell(row=dist_start, column=1)
    dc.value     = DIST_LABELS[dist]
    dc.font      = FONT_DIST
    dc.alignment = CENTER
    dc.border    = thin_border()

wb.save(OUTPUT_EXCEL)
print(f"\nSaved: {OUTPUT_EXCEL}")

if missing:
    print(f"\n{len(missing)} missing images:")
    for m in missing:
        print(f"  {m}")